from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

BASELINE_CENTER = "San Jose"
# Ruta relativa a la raiz del repo; funciona en local y en Streamlit Cloud
_HERE = Path(__file__).parent
DEFAULT_EXCEL = str(_HERE / "docs" / "Template V4b - San José (6).xlsx")


def norm(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if len(text) >= 2 and text[0] == "'" and text[-1] == "'":
        text = text[1:-1].strip()
    return text


def compact_openpyxl_sheet(file_path: Path, sheet_name: str, empty_streak_limit: int = 2500) -> pd.DataFrame:
    """Read sheet rows until a long empty streak appears.

    Several sheets in this workbook have formatted ranges up to ~1M rows.
    This avoids loading those empty tails while preserving real data rows.
    """
    wb = load_workbook(file_path, data_only=True, read_only=True)
    ws = wb[sheet_name]

    rows: list[list[str]] = []
    empty_streak = 0
    max_seen_cols = 0

    for row in ws.iter_rows(values_only=True):
        vals = [norm(v) for v in row]
        if any(vals):
            rows.append(vals)
            max_seen_cols = max(max_seen_cols, len(vals))
            empty_streak = 0
        else:
            if rows:
                empty_streak += 1
                if empty_streak >= empty_streak_limit:
                    break

    wb.close()

    if not rows:
        return pd.DataFrame()

    rows = [r + [""] * (max_seen_cols - len(r)) for r in rows]
    header = rows[0]
    data = rows[1:]

    safe_header: list[str] = []
    seen: dict[str, int] = {}
    for i, col in enumerate(header):
        col_name = norm(col) or f"col_{i+1}"
        if col_name in seen:
            seen[col_name] += 1
            col_name = f"{col_name}_{seen[col_name]}"
        else:
            seen[col_name] = 1
        safe_header.append(col_name)

    return pd.DataFrame(data, columns=safe_header)


@st.cache_data(show_spinner=False)
def load_data(file_path: Path) -> dict[str, pd.DataFrame]:
    convenios = compact_openpyxl_sheet(file_path, "Convenios")
    convenios_planes = compact_openpyxl_sheet(file_path, "Convenios_Planes")
    homologaciones = compact_openpyxl_sheet(file_path, "Homologación")
    prestaciones_catalogos = compact_openpyxl_sheet(file_path, "PrestacionesCatalogos")
    catalogos_convenio = compact_openpyxl_sheet(file_path, "CatalogosConvenio")
    modulos_reglas = compact_openpyxl_sheet(file_path, "ModulosReglas")
    reglas_moduladas_detalle = compact_openpyxl_sheet(file_path, "ReglasModuladasDetalle")
    clasif_prest_val_ref = compact_openpyxl_sheet(file_path, "ClasifPrestValoresReferencia")

    return {
        "convenios": convenios,
        "convenios_planes": convenios_planes,
        "homologaciones": homologaciones,
        "prestaciones_catalogos": prestaciones_catalogos,
        "catalogos_convenio": catalogos_convenio,
        "modulos_reglas": modulos_reglas,
        "reglas_moduladas_detalle": reglas_moduladas_detalle,
        "clasif_prest_val_ref": clasif_prest_val_ref,
    }


def merge_homologaciones_with_modulo(
    homologaciones: pd.DataFrame,
    prestaciones_catalogos: pd.DataFrame,
) -> pd.DataFrame:
    h = homologaciones.copy()
    p = prestaciones_catalogos.copy()

    needed_h = [
        "Catalogo",
        "codigo_referencia",
        "PrestacionReferencia",
        "Codigo_Catalogo",
        "PrestacionCatalogo",
        "Tipo Homologacion",
        "ESTADO",
        "fecha inicio vigencia",
    ]
    for col in needed_h:
        if col not in h.columns:
            h[col] = ""

    needed_p = ["Catalogo", "Codigo", "Es Modulo", "Modulo", "Modulo Orden", "ESTADO"]
    for col in needed_p:
        if col not in p.columns:
            p[col] = ""

    h["Catalogo"] = h["Catalogo"].map(norm)
    h["Codigo_Catalogo"] = h["Codigo_Catalogo"].map(norm)

    p["Catalogo"] = p["Catalogo"].map(norm)
    p["Codigo"] = p["Codigo"].map(norm)
    p["Es Modulo"] = p["Es Modulo"].map(norm)
    p["Modulo"] = p["Modulo"].map(norm)

    merged = h.merge(
        p[["Catalogo", "Codigo", "Es Modulo", "Modulo", "Modulo Orden"]],
        left_on=["Catalogo", "Codigo_Catalogo"],
        right_on=["Catalogo", "Codigo"],
        how="left",
    )

    merged["Es Modulo"] = merged["Es Modulo"].fillna("").map(norm)
    merged["Modulo"] = merged["Modulo"].fillna("").map(norm)

    merged.rename(
        columns={
            "codigo_referencia": "Codigo Referencia",
            "Codigo_Catalogo": "Codigo Catalogo",
            "PrestacionReferencia": "Prestacion Referencia",
            "PrestacionCatalogo": "Prestacion Catalogo",
            "Tipo Homologacion": "Tipo Homologacion",
            "ESTADO": "Estado",
            "fecha inicio vigencia": "Vigencia Inicio",
        },
        inplace=True,
    )

    return merged


def get_catalogos_for_convenio(catalogos_convenio: pd.DataFrame, convenio_name: str) -> list[str]:
    if catalogos_convenio.empty:
        return []

    df = catalogos_convenio.copy()
    if "Convenio" not in df.columns or "Catalogo" not in df.columns:
        return []

    mask = df["Convenio"].map(norm).str.upper() == convenio_name.upper()
    cats = sorted({norm(v) for v in df.loc[mask, "Catalogo"].tolist() if norm(v)})
    return cats


@st.cache_data(show_spinner=False)
def get_modulo_prestaciones_for_modulo(file_path: Path, modulo_name: str, empty_streak_limit: int = 5000) -> pd.DataFrame:
    wb = load_workbook(file_path, data_only=True, read_only=True)
    ws = wb["ModuloPrestacion"]

    rows: list[list[str]] = []
    empty_streak = 0
    headers: list[str] = []
    idx: dict[str, int] = {}

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = [norm(v) for v in row]
        if row_idx == 1:
            headers = vals
            idx = {h: i for i, h in enumerate(headers) if h}
            continue

        if not any(vals):
            empty_streak += 1
            if rows and empty_streak >= empty_streak_limit:
                break
            continue

        empty_streak = 0
        modulo_i = idx.get("Modulo", 0)
        modulo_val = vals[modulo_i] if modulo_i < len(vals) else ""
        if modulo_val.upper() != norm(modulo_name).upper():
            continue

        rows.append(vals)

    wb.close()

    if not rows:
        return pd.DataFrame(columns=["Modulo", "CodigoPrestacionReferencia", "PrestacionReferencia", "TipoInclusion", "Tope", "ESTADO"])

    out = pd.DataFrame(rows, columns=headers)
    for col in ["Modulo", "CodigoPrestacionReferencia", "PrestacionReferencia", "TipoInclusion", "Tope", "ESTADO"]:
        if col not in out.columns:
            out[col] = ""
        out[col] = out[col].map(norm)

    return out[["Modulo", "CodigoPrestacionReferencia", "PrestacionReferencia", "TipoInclusion", "Tope", "ESTADO"]]


def get_modulo_drilldown(file_path: Path, modulos_reglas: pd.DataFrame, modulo_name: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    mp = get_modulo_prestaciones_for_modulo(file_path, modulo_name).copy()
    mr = modulos_reglas.copy()

    for col in ["ReglaModulada", "TipoInclusion", "Orden", "Estado"]:
        if col not in mr.columns:
            mr[col] = ""

    mp_out = mp

    modulo_col_mr = "Modulo" if "Modulo" in mr.columns else mr.columns[0]
    mr_mask = mr[modulo_col_mr].map(norm).str.upper() == modulo_name.upper()
    mr_out = mr.loc[mr_mask, [modulo_col_mr, "ReglaModulada", "TipoInclusion", "Orden", "Estado"]]
    mr_out.rename(columns={modulo_col_mr: "Modulo"}, inplace=True)

    return mp_out.reset_index(drop=True), mr_out.reset_index(drop=True)


def get_regla_prestaciones_drilldown(
    regla_name: str,
    reglas_moduladas_detalle: pd.DataFrame,
    clasif_prest_val_ref: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    rmd = reglas_moduladas_detalle.copy()
    cpr = clasif_prest_val_ref.copy()

    for col in ["ReglaModulada", "Ambito", "ClasificacionValor", "ESTADO"]:
        if col not in rmd.columns:
            rmd[col] = ""
        rmd[col] = rmd[col].map(norm)

    for col in [
        "Nomenclador",
        "CodigoPrestacionNomencladorReferencia",
        "PrestacionNomencladorReferencia",
        "Clasificacion",
        "ClasificacionValor",
        "ESTADO",
    ]:
        if col not in cpr.columns:
            cpr[col] = ""
        cpr[col] = cpr[col].map(norm)

    rmd_view = rmd.loc[rmd["ReglaModulada"].str.upper() == norm(regla_name).upper(), ["ReglaModulada", "Ambito", "ClasificacionValor", "ESTADO"]].drop_duplicates()

    clasif_values = sorted([v for v in rmd_view["ClasificacionValor"].map(norm).unique().tolist() if v])
    # Fallback: algunas reglas pueden mapear directamente por el nombre de la regla.
    if not clasif_values and norm(regla_name):
        clasif_values = [norm(regla_name)]

    prest_view = cpr.loc[
        cpr["ClasificacionValor"].map(norm).str.upper().isin([v.upper() for v in clasif_values]),
        [
            "Nomenclador",
            "CodigoPrestacionNomencladorReferencia",
            "PrestacionNomencladorReferencia",
            "Clasificacion",
            "ClasificacionValor",
            "ESTADO",
        ],
    ].drop_duplicates()

    return rmd_view.reset_index(drop=True), prest_view.reset_index(drop=True)


@st.cache_data(show_spinner=False)
def load_osde_csv() -> pd.DataFrame:
    """Carga el CSV de OSDE desde docs/CONVENIO_BASE_OSDE/"""
    osde_folder = _HERE / "docs" / "CONVENIO_BASE_OSDE"
    csv_files = list(osde_folder.glob("*.csv"))
    
    if not csv_files:
        return pd.DataFrame()
    
    # Tomar el primer CSV (en caso de haber múltiples)
    csv_path = csv_files[0]
    
    try:
        df = pd.read_csv(csv_path, sep=";", encoding="utf-8-sig")
        return df
    except Exception:
        return pd.DataFrame()


def compare_osde_vs_template(osde_df: pd.DataFrame, template_codigos: set[str]) -> dict[str, Any]:
    """Compara IDs del CSV OSDE vs códigos del Template"""
    if osde_df.empty:
        return {"error": "No se pudo cargar el archivo OSDE"}
    
    # Extraer códigos de OSDE (columna B = índice 1, o la segunda columna)
    osde_codigos = set()
    if len(osde_df.columns) > 1:
        col_b = osde_df.columns[1]  # Segunda columna (índice 1)
        osde_codigos = {norm(str(v)) for v in osde_df[col_b].unique() if norm(str(v))}
    
    template_codigos_norm = {norm(c) for c in template_codigos if c}
    
    # Comparativa
    coincidencias = osde_codigos & template_codigos_norm
    solo_osde = osde_codigos - template_codigos_norm
    solo_template = template_codigos_norm - osde_codigos
    
    return {
        "osde_codigos": osde_codigos,
        "template_codigos": template_codigos_norm,
        "coincidencias": coincidencias,
        "solo_osde": solo_osde,
        "solo_template": solo_template,
        "osde_df": osde_df,
    }


def main() -> None:
    st.set_page_config(page_title="Baseline Convenios - San Jose", layout="wide")

    st.title("Baseline de Configuracion de Convenios")
    st.caption(f"Centro baseline: {BASELINE_CENTER}")

    excel_path = Path(st.text_input("Ruta Excel", value=DEFAULT_EXCEL)).expanduser()
    if not excel_path.exists():
        st.error("No se encontro el archivo Excel en la ruta indicada.")
        return

    with st.spinner("Cargando y normalizando hojas del template..."):
        data = load_data(excel_path)
    
    # Crear tabs: Template y OSDE
    tab_template, tab_osde = st.tabs(["Template San Jose", "OSDE Comparativa"])
    
    # ===== TAB TEMPLATE =====
    with tab_template:
        template_main(data, excel_path)
    
    # ===== TAB OSDE =====
    with tab_osde:
        osde_main(data, excel_path)


def template_main(data: dict[str, pd.DataFrame], excel_path: Path) -> None:
    """Contenido original de main() - pestaña Template"""
    
    convenios_planes = data["convenios_planes"]
    homologaciones = data["homologaciones"]
    prestaciones_catalogos = data["prestaciones_catalogos"]
    catalogos_convenio = data["catalogos_convenio"]
    modulos_reglas = data["modulos_reglas"]
    reglas_moduladas_detalle = data["reglas_moduladas_detalle"]
    clasif_prest_val_ref = data["clasif_prest_val_ref"]

    if catalogos_convenio.empty and prestaciones_catalogos.empty:
        st.error("No se pudieron leer datos de CatalogosConvenio/PrestacionesCatalogos.")
        return

    cc = catalogos_convenio.copy()
    cp = convenios_planes.copy()
    pc = prestaciones_catalogos.copy()

    for col in ["Catalogo", "Convenio", "ESTADO"]:
        if col not in cc.columns:
            cc[col] = ""
        cc[col] = cc[col].map(norm)

    for col in ["Convenio", "Financiador", "Plan", "Estado"]:
        if col not in cp.columns:
            cp[col] = ""
        cp[col] = cp[col].map(norm)

    for col in ["Catalogo", "Codigo", "Nombre", "Es Modulo", "Modulo", "Modulo Orden", "ESTADO"]:
        if col not in pc.columns:
            pc[col] = ""
        pc[col] = pc[col].map(norm)

    st.subheader("Filtros superiores")
    c1, c2, c3, c4, c5 = st.columns([1.2, 1.2, 1.2, 1.4, 1.0])

    catalogo_options = sorted({v for v in cc["Catalogo"].tolist() if v} | {v for v in pc["Catalogo"].tolist() if v})
    catalogo_sel = c1.selectbox("Catalogo", options=catalogo_options, index=0 if catalogo_options else None)

    convenios_catalogo = sorted(
        cc.loc[cc["Catalogo"].str.upper() == norm(catalogo_sel).upper(), "Convenio"].dropna().map(norm).unique().tolist()
    )
    convenio_options = ["(Todos)"] + [v for v in convenios_catalogo if v]
    convenio_sel = c2.selectbox("Convenio", options=convenio_options, index=0)

    cp_conv = cp.loc[cp["Convenio"].isin(convenios_catalogo)].copy()
    if convenio_sel != "(Todos)":
        cp_conv = cp_conv.loc[cp_conv["Convenio"].str.upper() == norm(convenio_sel).upper()]

    fin_options = ["(Todos)"] + sorted([v for v in cp_conv["Financiador"].unique().tolist() if v])
    fin_sel = c3.selectbox("Financiador asociado", options=fin_options, index=0)
    cp_fin = cp_conv.copy()
    if fin_sel != "(Todos)":
        cp_fin = cp_fin.loc[cp_fin["Financiador"].str.upper() == norm(fin_sel).upper()]

    plan_options = ["(Todos)"] + sorted([v for v in cp_fin["Plan"].unique().tolist() if v])
    plan_sel = c4.selectbox("Plan", options=plan_options, index=0)
    cp_view = cp_fin.copy()
    if plan_sel != "(Todos)":
        cp_view = cp_view.loc[cp_view["Plan"].str.upper() == norm(plan_sel).upper()]

    only_modulo = c5.checkbox("Solo modulos", value=False)

    st.markdown("---")

    # Context cards
    cards = st.columns(4)
    cards[0].metric("Centro", BASELINE_CENTER)
    cards[1].metric("Catalogo", norm(catalogo_sel) or "-")
    cards[2].metric("Convenio", norm(convenio_sel) or "(Todos)")
    cards[3].metric("Financiador / Plan", f"{norm(fin_sel) or '(Todos)'} / {norm(plan_sel) or '(Todos)'}")

    st.markdown("Asociaciones Convenio / Financiador / Plan para el catalogo seleccionado")
    st.dataframe(
        cp_view[["Convenio", "Financiador", "Plan", "Estado"]].drop_duplicates().reset_index(drop=True),
        width="stretch",
        height=180,
    )

    hbase = merge_homologaciones_with_modulo(homologaciones, prestaciones_catalogos)
    hbase["Catalogo"] = hbase["Catalogo"].map(norm)

    hview = hbase.loc[hbase["Catalogo"].str.upper() == norm(catalogo_sel).upper()].copy()
    pview = pc.loc[pc["Catalogo"].str.upper() == norm(catalogo_sel).upper()].copy()

    # Brecha: prestaciones del catalogo sin asociacion de homologacion (codigo catalogo -> codigo referencia)
    h_catalog = hbase.loc[hbase["Catalogo"].str.upper() == norm(catalogo_sel).upper()].copy()
    codigos_homologados = {
        norm(v)
        for v in h_catalog["Codigo Catalogo"].tolist()
        if norm(v)
    }
    p_no_homo = pview.loc[~pview["Codigo"].map(norm).isin(codigos_homologados)].copy()

    # Extra filters on top for readability
    f1, f2, f3 = st.columns([1.0, 1.0, 2.2])
    estado_opts = sorted([v for v in hview["Estado"].map(norm).unique().tolist() if v])
    tipo_h_opts = sorted([v for v in hview["Tipo Homologacion"].map(norm).unique().tolist() if v])

    estado_sel = f1.multiselect("Estado homologacion", options=estado_opts, default=estado_opts)
    tipo_h_sel = f2.multiselect("Tipo homologacion", options=tipo_h_opts, default=tipo_h_opts)
    search_txt = f3.text_input("Buscar por codigo o nombre", value="")

    if estado_sel:
        hview = hview.loc[hview["Estado"].map(norm).isin(estado_sel)]
    if tipo_h_sel:
        hview = hview.loc[hview["Tipo Homologacion"].map(norm).isin(tipo_h_sel)]
    if only_modulo:
        hview = hview.loc[hview["Es Modulo"].map(norm).str.upper() == "SI"]
        pview = pview.loc[pview["Es Modulo"].map(norm).str.upper() == "SI"]

    search_txt_norm = norm(search_txt).upper()
    if search_txt_norm:
        mask = (
            hview["Codigo Referencia"].map(norm).str.upper().str.contains(search_txt_norm, na=False)
            | hview["Codigo Catalogo"].map(norm).str.upper().str.contains(search_txt_norm, na=False)
            | hview["Prestacion Referencia"].map(norm).str.upper().str.contains(search_txt_norm, na=False)
            | hview["Prestacion Catalogo"].map(norm).str.upper().str.contains(search_txt_norm, na=False)
            | hview["Modulo"].map(norm).str.upper().str.contains(search_txt_norm, na=False)
        )
        hview = hview.loc[mask]
        pmask = (
            pview["Codigo"].map(norm).str.upper().str.contains(search_txt_norm, na=False)
            | pview["Nombre"].map(norm).str.upper().str.contains(search_txt_norm, na=False)
            | pview["Modulo"].map(norm).str.upper().str.contains(search_txt_norm, na=False)
        )
        pview = pview.loc[pmask]

    st.subheader("Homologaciones configuradas (baseline San Jose)")
    k1, k2, k3 = st.columns(3)
    k1.metric("Homologaciones visibles", len(hview))
    k2.metric("Prestaciones catalogo", len(pview))
    k3.metric("Filas modulo (SI)", int((pview["Es Modulo"].map(norm).str.upper() == "SI").sum()))

    cols_show = [
        "Catalogo",
        "Codigo Referencia",
        "Prestacion Referencia",
        "Codigo Catalogo",
        "Prestacion Catalogo",
        "Tipo Homologacion",
        "Es Modulo",
        "Modulo",
        "Estado",
        "Vigencia Inicio",
    ]
    st.dataframe(hview[cols_show], width="stretch", height=420)

    st.subheader("Prestaciones del catalogo")
    pcols_show = ["Catalogo", "Codigo", "Nombre", "Es Modulo", "Modulo", "Modulo Orden", "ESTADO"]
    st.dataframe(pview[pcols_show], width="stretch", height=360)

    st.subheader("Prestaciones del catalogo sin homologacion")
    st.caption("Codigos de catalogo que no tienen asociacion a ningun codigo de referencia en la hoja Homologación.")
    g1, g2 = st.columns(2)
    g1.metric("Prestaciones sin homologacion", len(p_no_homo))
    g2.metric("Prestaciones homologadas", max(len(pview) - len(p_no_homo), 0))
    st.dataframe(p_no_homo[pcols_show], width="stretch", height=260)

    st.subheader("Drilldown de modulos")
    modulos_visibles = sorted(
        [
            m
            for m in pview.loc[pview["Es Modulo"].map(norm).str.upper() == "SI", "Modulo"].map(norm).unique().tolist()
            if m
        ]
    )

    if not modulos_visibles:
        st.info("No hay prestaciones modulares en la seleccion actual.")
        return

    modulo_sel = st.selectbox("Seleccionar modulo", options=modulos_visibles, index=0)
    mp_out, mr_out = get_modulo_drilldown(excel_path, modulos_reglas, modulo_sel)

    if mr_out.empty:
        st.markdown("Prestaciones dentro del modulo")
        st.dataframe(mp_out, width="stretch", height=320)
        st.info("No hay reglas moduladas asociadas para el catalogo/modulo seleccionado.")
        return

    d1, d2 = st.columns(2)
    with d1:
        st.markdown("Prestaciones dentro del modulo")
        st.dataframe(mp_out, width="stretch", height=320)
    with d2:
        st.markdown("Reglas moduladas del modulo")
        st.dataframe(mr_out, width="stretch", height=320)

    st.subheader("Drilldown de regla modulada -> prestaciones asociadas")
    reglas_modulo = sorted([v for v in mr_out["ReglaModulada"].map(norm).unique().tolist() if v])
    reglas_globales = []
    if not reglas_moduladas_detalle.empty and "ReglaModulada" in reglas_moduladas_detalle.columns:
        reglas_globales = sorted([v for v in reglas_moduladas_detalle["ReglaModulada"].map(norm).unique().tolist() if v])

    mostrar_todas = st.checkbox("Mostrar todas las reglas moduladas", value=False)
    reglas_options = reglas_globales if mostrar_todas else reglas_modulo
    if not reglas_options:
        reglas_options = reglas_globales

    if not reglas_options:
        st.info("No se encontraron reglas moduladas para mostrar.")
        return

    regla_sel = st.selectbox("Regla modulada", options=reglas_options, index=0)
    rmd_view, prest_regla = get_regla_prestaciones_drilldown(
        regla_sel,
        reglas_moduladas_detalle,
        clasif_prest_val_ref,
    )

    rd1, rd2 = st.columns(2)
    with rd1:
        st.markdown("Detalle de asociacion de la regla")
        st.dataframe(rmd_view, width="stretch", height=280)
    with rd2:
        st.markdown("Prestaciones asociadas a la regla")
        st.dataframe(prest_regla, width="stretch", height=280)


def osde_main(data: dict[str, pd.DataFrame], excel_path: Path) -> None:
    """Pestaña OSDE - Comparativa de IDs vs Template"""
    
    st.subheader("Comparativa OSDE vs Template")
    st.caption("Compara códigos del archivo OSDE vs códigos del Template San Jose")
    
    # Cargar datos
    with st.spinner("Cargando archivo OSDE..."):
        osde_df = load_osde_csv()
    
    if osde_df.empty:
        st.error("No se pudo cargar el archivo OSDE. Verifica que exista en docs/CONVENIO_BASE_OSDE/")
        return
    
    # Extraer códigos del Template (Prestaciones del Catálogo)
    pc = data["prestaciones_catalogos"].copy()
    for col in ["Catalogo", "Codigo"]:
        if col not in pc.columns:
            pc[col] = ""
        pc[col] = pc[col].map(norm)
    
    template_codigos = set(pc["Codigo"].unique())
    template_codigos = {c for c in template_codigos if c}
    
    # Comparar
    comparison = compare_osde_vs_template(osde_df, template_codigos)
    
    if "error" in comparison:
        st.error(comparison["error"])
        return
    
    osde_codigos = comparison["osde_codigos"]
    template_codigos_norm = comparison["template_codigos"]
    coincidencias = comparison["coincidencias"]
    solo_osde = comparison["solo_osde"]
    solo_template = comparison["solo_template"]
    
    # Métricas
    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Códigos OSDE", len(osde_codigos))
    m2.metric("Códigos Template", len(template_codigos_norm))
    m3.metric("Coincidencias", len(coincidencias), delta=f"{len(coincidencias) / max(len(osde_codigos), 1) * 100:.1f}%")
    m4.metric("Solo en OSDE", len(solo_osde))
    m5.metric("Solo en Template", len(solo_template))
    
    st.markdown("---")
    
    # Crear DataFrame consolidado para visualización
    all_codigos = sorted(osde_codigos | template_codigos_norm)
    
    comparativa_data = []
    for codigo in all_codigos:
        en_osde = "✓" if codigo in osde_codigos else ""
        en_template = "✓" if codigo in template_codigos_norm else ""
        estado = "Coincide" if (codigo in osde_codigos and codigo in template_codigos_norm) else ("Solo OSDE" if codigo in solo_osde else "Solo Template")
        
        comparativa_data.append({
            "Codigo": codigo,
            "En OSDE": en_osde,
            "En Template": en_template,
            "Estado": estado,
        })
    
    comparativa_df = pd.DataFrame(comparativa_data)
    
    # Filtros
    f1, f2 = st.columns([1.0, 2.0])
    
    estado_filter = f1.multiselect(
        "Filtrar por estado",
        options=["Coincide", "Solo OSDE", "Solo Template"],
        default=["Coincide", "Solo OSDE", "Solo Template"]
    )
    
    search_codigo = f2.text_input("Buscar código", value="")
    
    # Aplicar filtros
    df_filtered = comparativa_df.copy()
    
    if estado_filter:
        df_filtered = df_filtered[df_filtered["Estado"].isin(estado_filter)]
    
    if search_codigo:
        search_norm = norm(search_codigo).upper()
        df_filtered = df_filtered[df_filtered["Codigo"].str.upper().str.contains(search_norm, na=False)]
    
    # Mostrar tabla
    st.subheader("Tabla de Doble Entrada - Coincidencias y Diferencias")
    st.dataframe(
        df_filtered,
        width="stretch",
        height=500,
        column_config={
            "Codigo": st.column_config.TextColumn("Código", width=150),
            "En OSDE": st.column_config.TextColumn("En OSDE", width=100),
            "En Template": st.column_config.TextColumn("En Template", width=100),
            "Estado": st.column_config.TextColumn("Estado", width=150),
        }
    )
    
    # Descarga a Excel
    st.markdown("---")
    st.subheader("Descargar Resultados")
    
    # Crear archivo Excel con hojas
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Hoja 1: Comparativa completa
        df_filtered.to_excel(writer, index=False, sheet_name="Comparativa")
        
        # Hoja 2: Solo coincidencias
        df_coincide = df_filtered[df_filtered["Estado"] == "Coincide"].copy()
        df_coincide.to_excel(writer, index=False, sheet_name="Coincidencias")
        
        # Hoja 3: Solo en OSDE
        df_osde_only = df_filtered[df_filtered["Estado"] == "Solo OSDE"].copy()
        df_osde_only.to_excel(writer, index=False, sheet_name="Solo OSDE")
        
        # Hoja 4: Solo en Template
        df_template_only = df_filtered[df_filtered["Estado"] == "Solo Template"].copy()
        df_template_only.to_excel(writer, index=False, sheet_name="Solo Template")
    
    output.seek(0)
    
    st.download_button(
        label="📥 Descargar Comparativa en Excel",
        data=output.getvalue(),
        file_name="OSDE_vs_Template_Comparativa.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    main()
